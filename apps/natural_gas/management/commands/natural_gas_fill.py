from datetime import datetime
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from pathlib import Path
from warnings import catch_warnings, filterwarnings

from ...models import NaturalGasUsage


class Command(BaseCommand):
    """
    Import natural gas usage data to the database, from Excel (.xlsx) file.
    """
    help = "Import natural gas usage data."

    def handle(self, *args, **options):

        try:
            # Find spreadsheets containing usage data, based on filename prefix.
            #   UsageData12142024.xlsx (UsageDataMMDDYYYY.xlsx)
            xlsx_prefix = settings.NATURAL_GAS_PREFIX
            spreadsheets = sorted(
                list(Path(".").glob(f"{xlsx_prefix}*.xlsx")),
                reverse=True
            )
            self.stdout.write(
                self.style.SUCCESS(
                    f"{len(spreadsheets)} spreadsheet(s) found."
                )
            )
            num_created = 0

            # Loop through each spreadsheet file found.
            for spreadsheet in spreadsheets:
                path = Path(spreadsheet)
                name = path.name

                # Filter warning about spreadsheet style.
                with catch_warnings():
                    filterwarnings(
                        action="ignore",
                        category=UserWarning,
                        module="openpyxl.styles.stylesheet"
                    )

                    # Load the spreadsheet workbook.
                    xlsx_wb = load_workbook(filename=name, read_only=True)
                    book_obj = xlsx_wb.active
                    sheet_obj = book_obj
                    title = sheet_obj.title or ""

                # Ensure that the title matches the prefix.
                assert title == xlsx_prefix, f"Invalid worksheet title {title}!"

                # Skip header to parse the following columns of data:
                #   Bill Month, Units Consumed (CCF), Period Start, Period End
                rows = list(sheet_obj.iter_rows())
                for row in rows[5:]:
                    row_month, row_value, row_start, row_end = row
                    row_ccf = float(row_value.value)

                    # Parse "Bill Month" column (i.e. "Mar, 2025") in each row.
                    row_dt = datetime.strptime(
                        row_month.value, "%b, %Y"
                    ).date()

                    # Map month to natural gas usage in CCF.
                    usage = {"month": row_dt, "ccf": row_ccf}

                    # Update or create the usage object in the database.
                    obj, created = NaturalGasUsage.objects.update_or_create(
                        **usage,
                        defaults=usage
                    )

                    # Show any newly created natural gas usage database objects.
                    if created:
                        num_created += 1
                        self.stdout.write(self.style.SUCCESS(
                            f"Created:\t{obj.month.strftime('%B %Y')}"
                            f" ({obj.month}) [{obj.ccf} CCF]"
                        ))

            # List count of newly created natural gas usage database objects.
            if num_created > 0:
                self.stdout.write(self.style.SUCCESS(
                    f"Total:\t\t{num_created}"
                ))
            self.stdout.write(self.style.SUCCESS("Done."))

        except Exception as exc:
            raise CommandError(exc)
