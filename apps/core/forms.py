from csv import DictReader
from datetime import date, datetime, time
from django.conf import settings
from django.core.validators import FileExtensionValidator
from django.forms import Form, FileField, ValidationError
from django.utils import timezone
from openpyxl import load_workbook
from pathlib import Path
from warnings import catch_warnings, filterwarnings


class UploadUsageDataForm(Form):
    """
    Form to handle upload of utility usage data file.
    """
    file = FileField(
        validators=(
            FileExtensionValidator(
                allowed_extensions=settings.USAGE_FILE_SUFFIXES
            ),
        )
    )

    def clean_file(self):
        """
        Validate various uploaded utility usage data files.
        """
        file = self.cleaned_data.get("file")
        if not file:
            raise ValidationError("Missing file!")

        encoding = "utf-8"
        file_path = Path(settings.MEDIA_ROOT, file.name)
        usage = []
        utility = None

        _valid_units = {
            "electric": ("TYPE", "Electric usage", "kWh"),
            "water": (" Units", " Gallons"),
        }

        # Get uploaded file name suffix, and validate MIME type.
        suffix = file_path.suffix.lower()

        # Write uploaded file locally.
        with open(file_path, mode="wb+") as fh_w:
            for chunk in file.chunks():
                fh_w.write(chunk)

        """
        Handle the uploaded file (which was just written) uniquely per utility.
        """

        """Electric or Water usage: comma-separated values (.csv)"""
        if suffix == ".csv":

            # Different encoding for electric usage CSV file.
            if file.name == settings.WATER_FILENAME:
                utility = "water"

            # Different encoding for electric usage CSV file.
            if file.name.startswith(settings.ELECTRIC_PREFIX):
                encoding += "-sig"
                utility = "electric"

            # Open CSV file that we just wrote, with proper encoding.
            with open(file_path, mode="r", encoding=encoding) as read_fh:
                csv_lines = read_fh.readlines()
                to_read = csv_lines

                if utility == "electric":

                    # Ensure electric usage (7th row, 5th column) is "kWh".
                    uf = csv_lines[6].split(",")[4]
                    unit_found = uf.split("(")[1].split(")")[0]
                    assert unit_found ==  _valid_units[utility][2], (
                        f"Invalid {utility} unit column!"
                    )

                    # Skip header of electric usage file.
                    to_read = csv_lines[6:]

                # Get valid unit for electric or water usage CSV rows.
                utility_unit = _valid_units[utility][1]

                # Read CSV of electric or water usage data, iterating rows.
                reader = DictReader(to_read)
                for row in reader:

                    # Electric or water CSV file rows must be valid units.
                    row_unit = row[_valid_units[utility][0]]
                    assert row_unit == utility_unit, (
                        f"Invalid {utility} unit ({row_unit})!"
                    )

                    """
                    Parse electric usage row.
                    """
                    if utility == "electric":

                        # Parse date/time columns for each electric usage row.
                        time_pcs = row["START TIME"].split(':')

                        # Map hour to electricity usage in floating point kWh.
                        usage.append({
                            "hour": timezone.make_aware(
                                value=datetime.combine(
                                    date=date.fromisoformat(row["DATE"]),
                                    time=time(
                                        hour=int(time_pcs[0]),
                                        minute=int(time_pcs[1])
                                    )
                                )
                            ),
                            "kwh": float(row["USAGE (kWh)"])
                        })

                    """
                    Parse water usage row.
                    """
                    if utility == "water":

                        # Parse date column for each water usage row.
                        date_pcs = row[" Time Interval"].strip().split('/')
                        date_iso = f"{date_pcs[2]}-{date_pcs[0]}-{date_pcs[1]}"

                        # Map each day to water usage floating point gallons.
                        usage.append({
                            "day": date.fromisoformat(date_iso),
                            "gallons": float(row[" Consumption"].strip())
                        })

        """Natural Gas usage: Microsoft Excel (.xlsx)"""
        xlsx_prefix = settings.NATURAL_GAS_PREFIX
        if suffix == ".xlsx" and file.name.startswith(xlsx_prefix):
            utility = "natural_gas"

            # Filter warnings about spreadsheet style.
            with catch_warnings():
                filterwarnings(
                    action="ignore",
                    category=UserWarning,
                    module="openpyxl.styles.stylesheet"
                )

                # Load/read the spreadsheet workbook.
                xlsx_wb = load_workbook(filename=file_path, read_only=True)
                book_obj = xlsx_wb.active
                sheet_obj = book_obj
                sheet_title = sheet_obj.title or ""

            # Confirm worksheet title.
            assert sheet_title == xlsx_prefix, (
                f"Invalid worksheet ({sheet_title})!"
            )

            # Gather all worksheet rows into a list.
            rows = list(sheet_obj.iter_rows())

            # Ensure natural gas unit is "CCF" (5th row, 2nd column).
            unit_found = rows[4][1].value.split('(')[1].split(')')[0]
            assert unit_found == "CCF", f"Invalid {utility} unit!"

            # Skip header rows to parse columns of natural gas usage data.
            #   Bill Month, Units Consumed (CCF), Period Start, Period End
            for row in rows[5:]:
                """
                Parse natural gas usage row.
                """
                row_month, row_ccf, row_start, row_end = row

                # Parse "Bill Month" column from each natural gas usage row.
                #    Example: "Mar, 2025"
                row_dt = datetime.strptime(
                    row_month.value, "%b, %Y"
                ).date()

                # Map each month to natural gas usage in CCF as floating point.
                usage.append({
                    "month": row_dt,
                    "ccf": float(row_ccf.value)
                })

        if usage:
            self.cleaned_data["usage"] = usage

        if utility:
            self.cleaned_data["utility"] = utility

        # Delete the local copy of the uploaded usage data file.
        if file_path and file_path.is_file() and file_path.exists():
            if file_path.unlink():
                return file
