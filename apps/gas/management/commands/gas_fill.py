from datetime import datetime
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from pathlib import Path
from ...models import GasUsage


class Command(BaseCommand):
    """
    Import natural gas usage data to the database, from Excel (.xlsx) file.
    """
    help = "Import natural gas usage data."

    def handle(self, *args, **options):

        try:
            # UsageData12142024.xlsx (UsageDataMMDDYYYY.xlsx)
            xlsx_prefix = settings.GAS_PREFIX
            spreadsheets = sorted(
                list(Path(".").glob(f"{xlsx_prefix}*.xlsx")),
                reverse=True
            )
            self.stdout.write(
                self.style.SUCCESS(
                    f"{len(spreadsheets)} spreadsheet(s) found."
                )
            )
            num_created = 0

            for spreadsheet in spreadsheets:
                path = Path(spreadsheet)
                name = path.name
                xlsx_wb = load_workbook(filename=name, read_only=True)
                book_obj = xlsx_wb.active
                sheet_obj = book_obj
                title = sheet_obj.title or ""
                assert title == xlsx_prefix, f"Invalid worksheet title {title}!"

                rows = list(sheet_obj.iter_rows())
                for row in rows[5:]:
                    row_month, row_value, row_start, row_end = row
                    row_ccf = float(row_value.value)
                    row_dt = datetime.strptime(
                        row_month.value, "%b, %Y"
                    ).date()

                    # Create dictionary mapping month to gas usage in CCF.
                    usage = {"month": row_dt, "ccf": row_ccf}

                    # Update or create the gas usage object in the database.
                    obj, created = GasUsage.objects.update_or_create(
                        **usage,
                        defaults=usage
                    )

                    # Show any newly created gas usage database objects.
                    if created:
                        num_created += 1
                        self.stdout.write(self.style.SUCCESS(
                            f"Created:\t{obj.month.strftime('%A %Y')}"
                            f" ({obj.month}) [{obj.ccf} CCF]"
                        ))

            # List total count of newly created gas usage database objects.
            if num_created > 0:
                self.stdout.write(self.style.SUCCESS(
                    f"Total:\t\t{num_created}"
                ))
            self.stdout.write(self.style.SUCCESS("Done."))

        except Exception as exc:
            raise CommandError(exc)
